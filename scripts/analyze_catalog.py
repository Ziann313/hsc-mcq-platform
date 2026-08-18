from collections import Counter
import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path("/home/ubuntu/upload/HSC_Bangladesh_Master_Curriculum_and_Book_Catalog_2025_26.xlsx")


def rows(sheet):
    values = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(cell is not None for cell in row)]


def cell(row, *names):
    for name in names:
        if name in row and row[name] is not None:
            return str(row[name]).strip()
    return ""


def quote(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def group_slug(value):
    return {
        "Common": "common",
        "Science": "science",
        "Humanities": "humanities",
        "Business Studies": "business-studies",
    }.get(value, "other")


def language_status(value):
    normalized = value.casefold()
    if normalized.startswith("yes"):
        return "verified"
    if normalized.startswith("no"):
        return "none"
    return "not_verified"


def main():
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    subject_rows = rows(workbook["Subject_Paper_Master"])
    chapter_rows = rows(workbook["Chapter_Master"])
    book_rows = rows(workbook["Book_Catalog"])
    source_rows = rows(workbook["Sources"])
    verification_rows = rows(workbook["Verification_Queue"])

    print("SHEET_COUNTS")
    for sheet in workbook.worksheets:
        print(f"{sheet.title}: rows={sheet.max_row} columns={sheet.max_column}")

    print("\nSUBJECT_SUMMARY")
    group_counts = Counter(cell(row, "Group") for row in subject_rows)
    for group, count in sorted(group_counts.items()):
        print(f"{group}: {count} paper rows")
    unique_subjects = {(cell(row, "Group"), cell(row, "Subject English")) for row in subject_rows}
    print(f"unique group-subject entries: {len(unique_subjects)}")
    print(f"chapter entries: {len(chapter_rows)}")

    print("\nBOOK_CATALOG")
    print("HEADERS=" + json.dumps(list(book_rows[0].keys()), ensure_ascii=False))
    for row in book_rows:
        print(json.dumps(row, ensure_ascii=False, default=str))

    print("\nSOURCES")
    for row in source_rows:
        print(" | ".join(f"{header}={value}" for header, value in row.items() if value is not None))

    print("\nVERIFICATION_QUEUE")
    for row in verification_rows:
        print(" | ".join(f"{header}={value}" for header, value in row.items() if value is not None))

    print("\nCATALOG_SUBJECT_IMPORT_SQL")
    values = []
    for row in subject_rows:
        values.append("(" + ", ".join([
            "1",
            "NULL",
            quote(group_slug(cell(row, "Group"))),
            quote(cell(row, "Subject Code")),
            quote(cell(row, "Paper")),
            quote(cell(row, "Subject English")),
            quote(cell(row, "Subject Bangla")),
            quote(language_status(cell(row, "English-Version Availability"))),
            "'needs_review'",
        ]) + ")")
    print("INSERT INTO catalog_subjects (academicYearId, subjectId, groupSlug, subjectCode, paperLabel, nameEn, nameBn, englishVersionAvailability, verificationStatus) VALUES\n" + ",\n".join(values) + ";")


if __name__ == "__main__":
    main()
